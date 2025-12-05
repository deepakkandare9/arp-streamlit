import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
import io


# ---------------------------------------------------------
# Extract IPv4 IPs
# ---------------------------------------------------------
def extract_ips(file):
    content = file.read().decode("utf-8", errors="ignore").splitlines()
    ips = []

    for ln in content:
        parts = ln.split()
        if parts:
            candidate = parts[0]
            if "." in candidate:
                seg = candidate.split(".")
                if len(seg) == 4 and all(s.isdigit() for s in seg):
                    ips.append(candidate)

    return pd.DataFrame({"IP": ips}).drop_duplicates()


# ---------------------------------------------------------
# Create Excel with Highlights
# ---------------------------------------------------------
def create_ip_comparison(pre_file, post_file):

    df_pre = extract_ips(pre_file)
    df_post = extract_ips(post_file)

    df_cmp = pd.merge(df_pre, df_post, on="IP", how="outer", indicator=True)
    df_cmp["Status"] = df_cmp["_merge"].map({
        "both": "Present in Pre & Post",
        "left_only": "Missing in Post",
        "right_only": "New in Post"
    })
    df_cmp = df_cmp[["IP", "Status"]]

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"ARP_IP_Comparison_{ts}.xlsx"

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_pre.to_excel(writer, sheet_name="Pre_IPs", index=False)
        df_post.to_excel(writer, sheet_name="Post_IPs", index=False)
        df_cmp.to_excel(writer, sheet_name="IP_Comparison", index=False)

    buffer.seek(0)
    wb = openpyxl.load_workbook(buffer)
    ws = wb["IP_Comparison"]

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    orange= PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        cell = row[0]
        if cell.value == "Present in Pre & Post":
            cell.fill = green
        elif cell.value == "Missing in Post":
            cell.fill = red
        elif cell.value == "New in Post":
            cell.fill = orange

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output, filename


# ---------------------------------------------------------
# STREAMLIT UI
# ---------------------------------------------------------
st.set_page_config(page_title="ARP Comparison Tool", layout="centered")

st.title("🔍 Cisco–Nokia ARP IP Comparison Tool")
st.write("Upload PRE & POST ARP logs → get Excel with color-coded comparison.")

pre_file = st.file_uploader("Upload PRE ARP File (Cisco)", type=["txt", "log"])
post_file = st.file_uploader("Upload POST ARP File (Nokia)", type=["txt", "log"])


if pre_file and post_file:
    if st.button("GENERATE EXCEL REPORT"):
        output, filename = create_ip_comparison(pre_file, post_file)

        st.success("Report ready! Download below 👇")
        st.download_button(
            label="⬇ Download Excel Report",
            data=output,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
